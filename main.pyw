from flask import Flask, request, send_file, jsonify, render_template, send_from_directory
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import io
import numpy as np
import threading
import socket
from code_py.func_app import check_heartbeat, start_executable, terminate_process, heartbeat
from code_py.config import json_add

def find_free_port():
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.bind(('localhost', 0))
    port = s.getsockname()[1]
    s.close()
    return port

app = Flask(__name__)

def rgb_to_hex(rgb):
    return '{:02x}{:02x}{:02x}'.format(*rgb)

@app.route('/')
def route_1():
    return render_template('index.html', json_add=json_add)

@app.route('/upload', methods=['POST'])
def upload_image():
    if 'image' not in request.files:
        return jsonify({'error': 'No image file provided'}), 400

    image_file = request.files['image']
    w = request.form['w']
    h = request.form['h']

    try:
        image = Image.open(image_file)
        if image.mode in ['L', 'LA', 'P', 'RGBA']:
            image = image.convert('RGB')

        new_width = int(w)
        new_height = int(h)
        image = image.resize((new_width, new_height), Image.LANCZOS)

        width, height = image.size

        wb = Workbook()
        ws = wb.active

        for y in range(height):
            for x in range(width):
                pixel_color = image.getpixel((x, y))
                if isinstance(pixel_color, int):
                    pixel_color = (pixel_color, pixel_color, pixel_color)
                hex_color = rgb_to_hex(pixel_color)
                fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                ws.cell(row=y+1, column=x+1).fill = fill

        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)

        return send_file(excel_io, as_attachment=True, download_name='image.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/preview', methods=['POST'])
def preview_image():
    if 'image' not in request.files:
        return jsonify({'error': 'No image file provided'}), 400

    image_file = request.files['image']
    w = request.form['w']
    h = request.form['h']

    try:
        image = Image.open(image_file)
        if image.mode in ['L', 'LA', 'P', 'RGBA']:
            image = image.convert('RGB')

        new_width = int(w)
        new_height = int(h)
        image = image.resize((new_width, new_height), Image.LANCZOS)

        width, height = image.size

        wb = Workbook()
        ws = wb.active

        for y in range(height):
            for x in range(width):
                pixel_color = image.getpixel((x, y))
                if isinstance(pixel_color, int):
                    pixel_color = (pixel_color, pixel_color, pixel_color)
                hex_color = rgb_to_hex(pixel_color)
                fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                ws.cell(row=y+1, column=x+1).fill = fill

        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)

        wb = load_workbook(excel_io)
        ws = wb.active

        image = Image.new('RGB', (width, height))
        pixels = np.array(image)
        for y in range(height):
            for x in range(width):
                fill_color = ws.cell(row=y+1, column=x+1).fill
                if fill_color and fill_color.start_color.index != '00000000':
                    hex_color = fill_color.start_color.index
                    color = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
                else:
                    color = (255, 255, 255)
                pixels[y, x] = color

        image = Image.fromarray(pixels)
        
        img_io = io.BytesIO()
        image.save(img_io, format='PNG')
        img_io.seek(0)

        return send_file(img_io, as_attachment=True, download_name='preview.png', mimetype='image/png')

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/heartbeat', methods=['POST'])
def post_10():
    return heartbeat()

if __name__ == '__main__':
    port = find_free_port()

    html_content = f'<style>iframe{{position: fixed;height: 100%;width: 100%;top: 0%;left: 0%;}}</style><iframe src="http://127.0.0.1:{port}/" frameborder="0"></iframe>'

    file_content = f'''name = Convert-PX
window_h = 800
window_w = 1000
html = "{html_content}"
    '''

    file_path = 'start.article'

    with open(file_path, 'w') as file:
        file.write(file_content)


    start_executable()
    threading.Thread(target=check_heartbeat, daemon=True).start()
    
    try:
        app.run(debug=True, port=port)
    finally:
        terminate_process()